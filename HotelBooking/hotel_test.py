import time

import openpyxl
import pytest
import os
from attr.setters import validate
from openpyxl.workbook import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.devtools.v85.web_audio import GraphObjectId
from selenium.webdriver.support.expected_conditions import frame_to_be_available_and_switch_to_it

from HotelBooking import Loginpage
from selenium import webdriver
from HotelBooking import HotelFramework as fw
from HotelBooking import SearchHotelPage as sp
from HotelBooking.BookAHotelPage import PAGENAME, fIRST_NAME, LAST_NAME, ADDRESS, CC_NUM, CC_TYPE, CC_DATE, CC_YEAR, \
    CVV, BOOKNOW_BTN, PROCESSINGTEXT
from HotelBooking import BookAHotelPage as bh
from HotelBooking.BookedItineraryPage import SEARCHBOX, CANCELBTN, GO, LOGOUT, BOOKEDITENERY, SEARCHRESULT, \
    BOOKEDITENERYBTN, LOGOUTMSG, CHECKALL
from HotelBooking.BookingconfirmationPage import ORDER_NO, BookingConfirmation, My_itinerary, \
    navigationToBookingConfirmationPage
from HotelBooking.HotelFramework import clickingtheElement
from HotelBooking.SearchHotelPage import SEARCHBTN, LOCATION, LOGOUTBTN
from HotelBooking.SelectHotelPage import USERNAME, HOTEL_RADIOBTN, CONTINUEBTN
from HotelBooking import SelectHotelPage
from openpyxl import workbook



@pytest.fixture #checking the login functionality of the Adactin Hotel
def login():
    global driver
    url = "https://adactinhotelapp.com/HotelAppBuild2/index.php"
    driver = fw.launchingBrowser("chrome")
    fw.navigateToUrl(driver, url)
    Loginpage.loginApplication(driver,"testerbee","Qwerty123$")
    driver.maximize_window()


def test_LoginSuccessful_01(login):#asserting title and url of the page after login

    Homepagetitle="Adactin.com - Search Hotel"
    assert Homepagetitle==driver.title
    searchpage_url="https://adactinhotelapp.com/HotelAppBuild2/SearchHotel.php"
    assert searchpage_url==driver.current_url


def test_validatemsg_02(login):#validating username message after login
    sp.welcomeMsgVerification(driver)


def test_searchHotelpage_03(login):#searching hotel based on preferences
    location=2
    hotelindex=2
    roomtype=2
    numOfRooms="1"
    datein="29/01/25"
    dateout="30/01/25"
    adultsno="2"
    childno="1"
    username="hello testerbee!"
    sp.typingRequirementforHotel(driver,location,hotelindex,roomtype,numOfRooms,datein,dateout,adultsno,childno)
    fw.verifying_elementText(driver,USERNAME,username)



def test_selectingHotel_04(login):#selecting the available hotel
    location = 2
    hotelindex = 2
    roomtype = 2
    numOfRooms = "1"
    datein = "29/01/25"
    dateout = "30/01/25"
    adultsno = "2"
    childno = "1"
    sp.typingRequirementforHotel(driver, location, hotelindex, roomtype, numOfRooms, datein, dateout, adultsno, childno)
    fw.clickingtheElement(driver,HOTEL_RADIOBTN)
    fw.clickingtheElement(driver,CONTINUEBTN)
    fw.isElementPresent(driver,PAGENAME)



def test_bookaHotel_05(login):#booking the hotel with credit card details and fetching the order id and storing it in the excel for next testcase
    location = 2
    hotelindex = 2
    roomtype = 2
    numOfRooms = "1"
    datein = "29/01/25"
    dateout = "30/01/25"
    adultsno = "2"
    childno = "1"
    sp.typingRequirementforHotel(driver, location, hotelindex, roomtype, numOfRooms, datein, dateout, adultsno, childno)
    fw.clickingtheElement(driver, HOTEL_RADIOBTN)
    fw.clickingtheElement(driver, CONTINUEBTN)
    firstname="priya"
    lastname="srinivasan"
    address="12th street,gandhipuram,Kelambakkam"
    creditCard="4321543265437654"
    ccType=1
    ccDate=2
    ccYear="2030"
    cvv="333"
    bh.BookingaHotel(driver,firstname,lastname,address,creditCard,ccType,ccDate,ccYear,cvv)
    time.sleep(10)
    fw.isElementPresent(driver, BookingConfirmation)
    fw.isElementPresent(driver, ORDER_NO)
    text=fw.gettingElementAttribute(driver, ORDER_NO)
    fw.clickingtheElement(driver, My_itinerary)
    wb = openpyxl.Workbook()
    sheet = wb.active
    c1 = sheet.cell(column=1, row=1)
    c1.value = text
    wb.save(os.path.join(os.path.dirname(__file__),"Excel","demo3.xlsx"))



#saved order id in the previous testcase is passed in the searchfield whether it is displayed.
def test_verifyingOrderno_06(login):#checking booking order id is found while searching in search fieldpu
    fw.clickingtheElement(driver,BOOKEDITENERYBTN)
    path= os.path.join(os.path.dirname(__file__),"Excel","demo3.xlsx")
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row=1, column=1)
    fw.passingData(driver, SEARCHBOX, cell_obj.value)
    fw.clickingtheElement(driver, GO)
    fw.isElementPresent(driver,SEARCHRESULT)
    fw.clickingtheElement(driver, CANCELBTN)
    fw.alertOk(driver)
    fw.isElementPresent(driver, SEARCHRESULT)


def test_deletingallbooking_07(login):#checking whether clicking it checkall button deletes all the bookings
    fw.clickingtheElement(driver,BOOKEDITENERYBTN)
    fw.clickingtheElement(driver,CHECKALL)
    fw.clickingtheElement(driver,CANCELBTN)
    fw.alertOk(driver)

def test_loggingout_08(login):#checking logout functionality
    fw.clickingtheElement(driver, LOGOUTBTN)
    fw.isElementPresent(driver, LOGOUTMSG)












